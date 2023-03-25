import datetime
import time
from datetime import datetime
from openpyxl import load_workbook
import openpyxl as opx

now = datetime.now()
response_date = now.strftime("%d.%m.%Y %H:%M:%S")



name = 'None'
surname = 'None'
day = '17'
year = '2005'
month = '04'


birthday = day + "." + str(month).replace(" ", "") + "." + year

filename = 'example.xlsx'

wb = load_workbook(filename)
ws = wb['Лист1']
ws.append([name, surname, birthday])
if(now.year - int(year)<18):
    ws.cell(row = ws.max_row, column = 3).font = opx.styles.Font(color='ff0816')
if(now.year - int(year)==18):
	if month=='04':
		if(int(day)>17):
			ws.cell(row = ws.max_row, column = 3).font = opx.styles.Font(color='ff0816')   

if(now.year - int(year)==18):
	if month == '05' or month == '06' or month == '07' or month == '08' or month == '09' or month == '10' or month == '11' or month == '12':
		ws.cell(row = ws.max_row, column = 3).font = opx.styles.Font(color='ff0816')


   

print(ws.max_row)
wb.save(filename)
print("saved 1")
wb.close()


input('wdw')

